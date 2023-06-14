import openpyxl


def separador():
    print('-='*100)


def apresentacao():
    print('''Bem-vindo ao automatizador de planilhas
com ele é possivel criar planilhas com base nos dados desejados
para começar vamos criar uma nova página dentrp de uma planilha
''')


class planilha_automatica:
    def __init__(self):
        self.continuar = None
        self.cabecalhos = None
        self.sheet_atual = None
        self.wb = openpyxl.Workbook()

    def criando_paginas(self):
        nome = ''
        while True:
            separador()
            nome = input('digite o nome da pagina desejada: ')
            separador()
            if len(nome) > 0:
                self.wb.create_sheet(nome)
                break
            else:
                print('por favor, digite o nome da pagina!')
                continue
        while True:
            separador()
            criar_sheets = input('deseja criar mais uma pagina? (sim/nao) ').lower().strip()
            separador()
            if criar_sheets == 'sim':
                separador()
                nome = input('digite o nome da pagina desejada: ')
                separador()
                if len(nome) > 0:
                    try:
                        del self.wb['Sheet']
                    except:
                        pass
                    self.wb.create_sheet(nome)
                else:
                    print('por favor informe um nome!')
            elif criar_sheets == 'nao':
                try:
                    del self.wb['Sheet']
                except:
                    pass
                break
            else:
                print('por favor responda sim ou nao')

    def escolher_sheet(self):
        print(self.wb.sheetnames)
        while True:
            if len(self.wb.sheetnames) > 1:
                try:
                    separador()
                    sheet = input('digite o nome da planilha na qual serão inseridos dados: ')
                    separador()
                    if sheet not in self.wb.sheetnames:
                        print('valor invalido, por favor digite novamente')
                    else:
                        self.sheet_atual = self.wb[sheet]
                        break
                except:
                    print('valor invalido, por favor digite novamente!')
            else:
                self.sheet_atual = self.wb[self.wb.sheetnames[0]]
                break
        separador()
        print(f'serão inseridos dados na {self.sheet_atual}')
        separador()

    def inserir_colunas(self):
        self.cabecalhos = []
        separador()
        cabecalho = input('digite o nome do primeiro cabecalho ou enter para deixar vazio: ')
        separador()
        self.cabecalhos.append(cabecalho)
        while True:
            separador()
            continuar = input('deseja adicionar mais colunas? (sim/nao) ')
            separador()
            if continuar == 'sim':
                separador()
                cabecalho = input('digite o nome do cabeçalho: ')
                separador()
                self.cabecalhos.append(cabecalho)
            elif continuar == 'nao':
                break
            else:
                print('por favor digite apenas sim ou nao!')
        self.sheet_atual.append(self.cabecalhos)

    def inserir_conteudo(self):
        inserir_dados = input('deseja inserir dados na planilha: (sim/nao)')
        if inserir_dados == 'sim':
            linhas = []
            while True:
                linhas.clear()
                for cabecalho in self.cabecalhos:
                    separador()
                    conteudo = input(f'insira conteudo para a coluna {cabecalho}: ')
                    separador()
                    linhas.append(conteudo)
                self.sheet_atual.append(linhas)
                while True:
                    separador()
                    continuar = input('deseja continuar? (sim/nao): ')
                    separador()
                    if continuar != 'sim' and continuar != 'nao':
                        separador()
                        print('por favor digite apenas sim ou nao!')
                        separador()
                    else:
                        break
                if continuar == 'nao':
                    break

    def salvar_wb(self):
        separador()
        nome = input('digite o nome da sua planilha : ')
        separador()
        self.wb.save(f'{nome}.xlsx')
        print('planilha salva com sucesso!')

